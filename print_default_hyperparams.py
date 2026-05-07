# -*- coding: utf-8 -*-
"""
打印 train_hybrid.py 的默认超参数，并保存为 xlsx 文件。
依赖：openpyxl（pip install openpyxl）
"""
import argparse
from pathlib import Path

# Default hyperparameters (aligned with train_hybrid.py)
DEFAULT_HYPERPARAMS = [
    ("data_root", "data", "Data root directory"),
    ("seq_len", 128, "Sequence length (time steps)"),
    ("batch_size", 64, "Transformer batch size"),
    ("epochs", 50, "Max Transformer training epochs"),
    ("lr", 0.001, "Transformer learning rate"),
    ("d_model", 64, "Transformer hidden dimension"),
    ("nhead", 4, "Number of attention heads"),
    ("num_layers", 2, "Transformer encoder layers"),
    ("dim_ff", 256, "Feed-forward dimension"),
    ("dropout", 0.1, "Dropout rate"),
    ("pool", "last", "Pooling: last | gap"),
    ("patience", 10, "Early stop: stop after N epochs without val improvement"),
    ("slide_stride", 64, "Sliding window stride for long trajectories"),
    ("num_segments", 3, "Number of segments per trajectory (head/mid/tail)"),
    ("segment_agg", "mean", "Segment aggregation: mean | max"),
    ("save_transformer", "checkpoints/feature_transformer.pt", "Transformer checkpoint path"),
    ("lgb_rounds", 500, "Max LightGBM boosting rounds"),
    ("lgb_early_stop", 50, "LightGBM early stopping rounds"),
    ("lgb_device", "cpu", "LightGBM device: cpu | gpu"),
    ("no_cuda", False, "Force PyTorch to use CPU"),
    ("use_optuna", False, "Use Optuna to search LightGBM hyperparameters"),
    ("optuna_trials", 30, "Number of Optuna trials"),
    ("no_repair", False, "Disable AIS trajectory interpolation repair"),
]


def print_defaults():
    """Print default hyperparameters table to terminal."""
    print("=" * 70)
    print("train_hybrid.py Default Hyperparameters")
    print("=" * 70)
    print(f"  {'Parameter':<24} {'Default':<28} {'Description'}")
    print("-" * 70)
    for name, value, desc in DEFAULT_HYPERPARAMS:
        val_str = str(value)
        if len(val_str) > 26:
            val_str = val_str[:23] + "..."
        print(f"  {name:<24} {val_str:<28} {desc}")
    print("=" * 70)


def save_xlsx(out_path: Path):
    """Write default hyperparameters to xlsx."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
    except ImportError:
        raise ImportError("Please install openpyxl: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Default Hyperparameters"

    # Table headers
    headers = ["Parameter", "Default", "Description"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 22

    # Data rows
    for row, (name, value, desc) in enumerate(DEFAULT_HYPERPARAMS, 2):
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=3, value=desc)

    # Column width
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 36

    wb.save(str(out_path))
    print(f"Saved: {out_path}")


def main():
    parser = argparse.ArgumentParser(description="Print default hyperparameters and save to xlsx")
    parser.add_argument("--out", type=str, default="default_hyperparams.xlsx", help="Output xlsx path")
    parser.add_argument("--no_print", action="store_true", help="Do not print to terminal, only generate xlsx")
    args = parser.parse_args()

    base_dir = Path(__file__).parent
    out_path = base_dir / args.out

    if not args.no_print:
        print_defaults()
    save_xlsx(out_path)


if __name__ == "__main__":
    main()
